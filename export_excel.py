"""Generate Excel workbook from a solved schedule."""
import io
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XlImage
from openpyxl.utils.units import cm_to_EMU, pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D

FILL_HEADER = PatternFill("solid", fgColor="1F4E79")
FILL_KLASICKA = PatternFill("solid", fgColor="BDD7EE")
FILL_FREESTYLE = PatternFill("solid", fgColor="9BC2E6")
FILL_SPORT = PatternFill("solid", fgColor="C6EFCE")
FILL_TEST = PatternFill("solid", fgColor="FFE699")
FILL_PRESUN = PatternFill("solid", fgColor="E2EFDA")

FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
FONT_NORMAL = Font(name="Calibri", size=11)
FONT_BOLD = Font(name="Calibri", size=11, bold=True)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

CATEGORY_FILL_BASE = {
    "klas": FILL_KLASICKA, "free": FILL_FREESTYLE,
    "sport": FILL_SPORT, "test": FILL_TEST, "presun": FILL_PRESUN,
}

ROW_HEIGHT_ACTIVITY = 30
ROW_HEIGHT_HEADER = 20

def time_to_min(s):
    h, m = s.split(":")
    return int(h) * 60 + int(m)


def min_to_time(m):
    return f"{m // 60:02d}:{m % 60:02d}"


def add_transfers(schedule, shared_names, station_lookup):
    result = []
    for i, (act, start, end, cat) in enumerate(schedule):
        result.append((act, start, end, cat))
        if i < len(schedule) - 1:
            next_act, next_start, _, _ = schedule[i + 1]
            curr_st = station_lookup(act, shared_names)
            next_st = station_lookup(next_act, shared_names)
            if (curr_st and next_st and curr_st != next_st
                    and curr_st != "shared" and next_st != "shared"):
                gap = time_to_min(next_start) - time_to_min(end)
                if gap >= 10:
                    result.append(("Presun", end, min_to_time(time_to_min(end) + 10), "presun"))
    return result


def generate_excel(teams, logo_path=None, competition_start="08:45",
                   competition_end="13:45", shared_events=None, num_teams=None,
                   config=None):
    """Generate Excel bytes from teams dict. Returns bytes."""
    if shared_events is None:
        shared_events = []
    if num_teams is None:
        num_teams = len(teams)

    comp_start_min = time_to_min(competition_start)
    comp_end_min = time_to_min(competition_end)

    shared_names = {ev.name for ev in shared_events}
    during_events = sorted(
        [ev for ev in shared_events if ev.overlaps_window(comp_start_min, comp_end_min)],
        key=lambda e: e.start_time)
    pre_events = sorted(
        [ev for ev in shared_events if ev.end_time <= comp_start_min],
        key=lambda e: e.start_time)
    post_events = sorted(
        [ev for ev in shared_events if ev.start_time >= comp_end_min],
        key=lambda e: e.start_time)

    category_fill = dict(CATEGORY_FILL_BASE)
    for ev in shared_events:
        category_fill[ev.category] = PatternFill("solid", fgColor=ev.color_bg.lstrip("#"))

    def station_lookup(act_name, sn):
        if config:
            st = config.get_station(act_name)
            if st:
                return st
        if act_name in sn:
            return "shared"
        return None

    if config:
        from solver import Activity
        mas_acts = config.masaze_activities
        test_act = config.test_activities[0]
    else:
        from solver import Activity
        mas_acts = [Activity("Klasická masáž", 20, "Klas.", "klas"),
                    Activity("Freestyle masáž", 20, "Free.", "free")]
        test_act = Activity("Test", 15, "Test", "test")

    teams_with_transfers = {t: add_transfers(teams[t], shared_names, station_lookup) for t in teams}

    wb = openpyxl.Workbook()

    # ═══ OVERVIEW SHEET ═══
    ws = wb.active
    ws.title = "Prehľad"

    ws.merge_cells("A1:M1")
    c = ws["A1"]
    c.value = "MASÉR V AKCII – Časový harmonogram"
    c.font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
    c.alignment = ALIGN_CENTER

    ws.merge_cells("A3:M3")
    ws["A3"].value = "Celkový program dňa"
    ws["A3"].font = Font(name="Calibri", size=13, bold=True)
    ws["A3"].alignment = ALIGN_LEFT

    all_display_events = sorted(pre_events + during_events + post_events,
                                key=lambda e: e.start_time)
    general_events = []
    for ev in all_display_events:
        if ev.num_groups > 1:
            sizes = ev.effective_group_sizes(num_teams)
            cumul = 0
            for g in range(ev.num_groups):
                gs = ev.group_starts[g]
                start_team = cumul + 1
                cumul += sizes[g]
                end_team = cumul
                general_events.append((
                    f"{min_to_time(gs)} – {min_to_time(gs + ev.duration)}",
                    f"{ev.name} – skupiny {start_team} až {end_team}"))
        else:
            general_events.append((
                f"{min_to_time(ev.start_time)} – {min_to_time(ev.end_time)}", ev.name))

    row = 4
    for time_str, desc in general_events:
        ws.cell(row=row, column=1, value=time_str).font = FONT_BOLD
        ws.cell(row=row, column=1).alignment = ALIGN_LEFT
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=13)
        ws.cell(row=row, column=2, value=desc).font = FONT_NORMAL
        ws.cell(row=row, column=2).alignment = ALIGN_LEFT
        ws.cell(row=row, column=2).border = THIN_BORDER
        for col in range(3, 14):
            ws.cell(row=row, column=col).border = THIN_BORDER
        row += 1

    row += 1

    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    ws.cell(row=row, column=1, value="Legenda farieb:").font = FONT_BOLD
    row += 1
    legend_items = []
    for a in mas_acts:
        legend_items.append((category_fill.get(a.category, FILL_KLASICKA), a.name))
    legend_items.append((FILL_SPORT, "Športové disciplíny (telocvičňa ZŠ)"))
    legend_items.append((category_fill.get(test_act.category, FILL_TEST), test_act.name))
    legend_items.append((FILL_PRESUN, "Presun medzi stanovišťami (~10 min)"))
    for ev in shared_events:
        legend_items.append((category_fill[ev.category], ev.name))
    for fill, label in legend_items:
        ws.cell(row=row, column=1).fill = fill
        ws.cell(row=row, column=1).border = THIN_BORDER
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=5)
        ws.cell(row=row, column=2, value=label).font = FONT_NORMAL
        ws.cell(row=row, column=2).border = THIN_BORDER
        for c2 in range(3, 6):
            ws.cell(row=row, column=c2).border = THIN_BORDER
        row += 1

    row += 1

    # Summary table
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=13)
    ws.cell(row=row, column=1, value="Prehľad stanovíšť podľa tímov").font = Font(
        name="Calibri", size=13, bold=True)
    row += 1

    sum_headers = ["Tím"] + [a.name for a in mas_acts]
    sport_col = len(sum_headers) + 1
    sum_headers.append("Športové disciplíny (telocvičňa ZŠ)")
    test_col = len(sum_headers) + 1
    sum_headers.append(test_act.name)
    for ev in during_events:
        sum_headers.append(ev.name)
    num_sum_cols = len(sum_headers)

    for ci, h in enumerate(sum_headers, 1):
        cell = ws.cell(row=row, column=ci, value=h)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    row += 1

    for t_id in sorted(teams.keys()):
        cell = ws.cell(row=row, column=1, value=f"Team {t_id}")
        cell.font = FONT_BOLD
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER

        sport_entries = [(s, e) for name, s, e, c in teams[t_id] if c == "sport"]
        if sport_entries:
            c2 = ws.cell(row=row, column=sport_col,
                         value=f"{sport_entries[0][0]} – {sport_entries[-1][1]}")
            c2.font = FONT_NORMAL
            c2.alignment = ALIGN_CENTER
            c2.border = THIN_BORDER
            c2.fill = FILL_SPORT

        col_map = {}
        for ci, a in enumerate(mas_acts, 2):
            col_map[a.name] = (ci, category_fill.get(a.category, FILL_KLASICKA))
        col_map[test_act.name] = (test_col, category_fill.get(test_act.category, FILL_TEST))
        for idx, ev in enumerate(during_events):
            col_map[ev.name] = (test_col + 1 + idx, category_fill[ev.category])

        for act_name, start, end, cat in teams[t_id]:
            if act_name in col_map:
                col, fill = col_map[act_name]
                c2 = ws.cell(row=row, column=col, value=f"{start} – {end}")
                c2.font = FONT_NORMAL
                c2.alignment = ALIGN_CENTER
                c2.border = THIN_BORDER
                c2.fill = fill

        for ci in range(1, num_sum_cols + 1):
            ws.cell(row=row, column=ci).border = THIN_BORDER
        row += 1

    row += 1

    # Timeline grid
    num_team_cols = len(teams)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_team_cols + 1)
    ws.cell(row=row, column=1, value="Časová os (5-minútové intervaly)").font = Font(
        name="Calibri", size=13, bold=True)
    row += 1

    cell = ws.cell(row=row, column=1, value="Čas")
    cell.font = FONT_HEADER
    cell.fill = FILL_HEADER
    cell.alignment = ALIGN_CENTER
    cell.border = THIN_BORDER
    for i, t_id in enumerate(sorted(teams.keys()), 1):
        cell = ws.cell(row=row, column=i + 1, value=f"Team {t_id}")
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    row += 1

    team_timeline = {}
    for t_id in teams:
        timeline = {}
        for act_name, start, end, cat in teams_with_transfers[t_id]:
            s, e = time_to_min(start), time_to_min(end)
            short = config.get_abbreviation(act_name) if config else act_name
            for m in range(s, e):
                timeline[m] = (short, cat)
        team_timeline[t_id] = timeline

    all_times = [comp_start_min, comp_end_min]
    for ev in shared_events:
        if ev.num_groups > 1:
            all_times.extend(ev.group_starts)
            all_times.extend([gs + ev.duration for gs in ev.group_starts])
        else:
            all_times.extend([ev.start_time, ev.end_time])
    tl_start = min(all_times)
    tl_end = max(all_times)

    for slot in range(tl_start, tl_end, 5):
        cell = ws.cell(row=row, column=1, value=min_to_time(slot))
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
        for i, t_id in enumerate(sorted(teams.keys()), 1):
            tl = team_timeline[t_id]
            act, cat = None, None
            for m in range(slot, min(slot + 5, tl_end)):
                if m in tl:
                    act, cat = tl[m]
                    break
            cell = ws.cell(row=row, column=i + 1)
            cell.border = THIN_BORDER
            cell.alignment = ALIGN_CENTER
            cell.font = Font(name="Calibri", size=9)
            if act:
                cell.value = act
                cell.fill = category_fill.get(cat, PatternFill())
        row += 1

    ws.column_dimensions["A"].width = 16
    for ci in range(2, num_team_cols + 2):
        ws.column_dimensions[get_column_letter(ci)].width = 14

    # ═══ PER-TEAM SHEETS ═══
    has_logo = logo_path and os.path.exists(logo_path)

    for t_id in sorted(teams.keys()):
        ws_t = wb.create_sheet(title=f"Team {t_id}")

        if has_logo:
            ws_t.row_dimensions[1].height = 60
            ws_t.row_dimensions[2].height = 6
            img = XlImage(logo_path)
            marker = AnchorMarker(col=0, row=0, colOff=cm_to_EMU(0.3), rowOff=cm_to_EMU(0.1))
            size = XDRPositiveSize2D(pixels_to_EMU(280), pixels_to_EMU(75))
            img.anchor = OneCellAnchor(_from=marker, ext=size)
            ws_t.add_image(img)
            title_row = 4
        else:
            title_row = 1

        ws_t.merge_cells(f"A{title_row}:D{title_row}")
        c = ws_t.cell(row=title_row, column=1)
        c.value = "MASÉR V AKCII"
        c.font = Font(name="Calibri", size=18, bold=True, color="1F4E79")
        c.alignment = ALIGN_CENTER
        ws_t.row_dimensions[title_row].height = 30

        info_row = title_row + 2
        ws_t.row_dimensions[title_row + 1].height = 6

        ws_t.cell(row=info_row, column=1,
                  value=f"Družstvo č. {t_id}").font = Font(name="Calibri", size=14, bold=True)
        ws_t.cell(row=info_row, column=1).alignment = ALIGN_LEFT
        ws_t.cell(row=info_row, column=1).border = THIN_BORDER

        ws_t.merge_cells(f"B{info_row}:D{info_row}")
        ws_t.cell(row=info_row, column=2,
                  value="Škola:").font = Font(name="Calibri", size=14, bold=True)
        ws_t.cell(row=info_row, column=2).alignment = ALIGN_LEFT
        ws_t.cell(row=info_row, column=2).border = THIN_BORDER
        ws_t.cell(row=info_row, column=3).border = THIN_BORDER
        ws_t.cell(row=info_row, column=4).border = THIN_BORDER
        ws_t.row_dimensions[info_row].height = 25

        hdr_row = info_row + 2
        for ci, hdr in enumerate(["Aktivita", "Čas", "Poznámka", "Body"], 1):
            cell = ws_t.cell(row=hdr_row, column=ci, value=hdr)
            cell.font = FONT_HEADER
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        ws_t.row_dimensions[hdr_row].height = ROW_HEIGHT_HEADER
        r = hdr_row + 1

        # Build group notes for grouped shared events
        group_notes = {}
        for ev in during_events:
            if ev.num_groups > 1:
                sizes = ev.effective_group_sizes(num_teams)
                cumul = 0
                for g in range(ev.num_groups):
                    start_team = cumul + 1
                    cumul += sizes[g]
                    end_team = cumul
                    if t_id >= start_team and t_id <= end_team:
                        group_notes[ev.name] = (
                            f"{g+1}. skupina (tímy {start_team}–{end_team})")
                        break

        for act_name, start, end, cat in teams_with_transfers[t_id]:
            fill = category_fill.get(cat, PatternFill())
            ws_t.cell(row=r, column=1, value=act_name).font = FONT_NORMAL
            ws_t.cell(row=r, column=1).alignment = ALIGN_LEFT
            ws_t.cell(row=r, column=1).border = THIN_BORDER
            ws_t.cell(row=r, column=1).fill = fill

            ws_t.cell(row=r, column=2, value=f"{start} – {end}").font = FONT_NORMAL
            ws_t.cell(row=r, column=2).alignment = ALIGN_CENTER
            ws_t.cell(row=r, column=2).border = THIN_BORDER
            ws_t.cell(row=r, column=2).fill = fill

            note = group_notes.get(act_name, "")
            ws_t.cell(row=r, column=3, value=note).font = FONT_NORMAL
            ws_t.cell(row=r, column=3).alignment = ALIGN_LEFT
            ws_t.cell(row=r, column=3).border = THIN_BORDER

            ws_t.cell(row=r, column=4).border = THIN_BORDER
            ws_t.cell(row=r, column=4).alignment = ALIGN_CENTER

            ws_t.row_dimensions[r].height = ROW_HEIGHT_ACTIVITY
            r += 1

        ws_t.column_dimensions["A"].width = 28
        ws_t.column_dimensions["B"].width = 18
        ws_t.column_dimensions["C"].width = 24
        ws_t.column_dimensions["D"].width = 12

        ws_t.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
            fitToPage=True)
        ws_t.page_setup.fitToWidth = 1
        ws_t.page_setup.fitToHeight = 1
        ws_t.page_setup.orientation = "portrait"
        ws_t.page_setup.paperSize = ws_t.PAPERSIZE_A4
        ws_t.page_margins.left = 0.5
        ws_t.page_margins.right = 0.5
        ws_t.page_margins.top = 0.4
        ws_t.page_margins.bottom = 0.4

    # ═══ SPORT SCORING SHEETS ═══
    SCHOOLS = [
        "SZŠ B. Bystrica",
        "SZŠ Košice",
        "SZŠ Lipt. Mikuláš",
        "SZŠ Michalovce",
        "SZŠ Nové Zámky",
        "SZŠ P. Bystrica",
        "SZŠ Prešov",
        "SZŠ Rožňava",
        "SZŠ Skalica",
        "SZŠ Š. Kluberta Levoča",
        "SZŠ Žilina",
    ]

    SPORT_SHEETS = [
        {
            "title": "Frisbee na cieľ",
            "members": 2,
            "max_per_member": 5,
            "columns": ["Pokus 1", "Pokus 2", "Pokus 3", "Pokus 4", "Pokus 5", "Spolu"],
            "legend": [
                "Každý člen tímu má 5 pokusov.",
                "Zásah cieľa = 1 bod, mimo = 0 bodov.",
                "Max. 5 bodov na člena, max. 10 bodov za tím.",
            ],
            "scoring_table": None,
        },
        {
            "title": "Beh na 50m",
            "members": 2,
            "max_per_member": 5,
            "columns": ["Čas (s)", "Body"],
            "legend": ["Čím kratší čas, tým viac bodov. Max. 5 bodov na člena, max. 10 za tím."],
            "scoring_table": {
                "headers": ["Body", "1b", "2b", "3b", "4b", "5b"],
                "rows": [
                    ["Chlapci", "9,0s a viac", "8,6–8,9s", "8,1–8,5s", "7,7–8,0s", "7,6s a menej"],
                    ["Dievčatá", "10,0s a viac", "9,5–9,9s", "8,9–9,4s", "8,4–8,8s", "8,3s a menej"],
                ],
            },
        },
        {
            "title": "Hod medicinbalom",
            "members": 2,
            "max_per_member": 5,
            "columns": ["Vzdialenosť (m)", "Body"],
            "legend": ["Čím dlhší hod, tým viac bodov. Max. 5 bodov na člena, max. 10 za tím."],
            "scoring_table": {
                "headers": ["Body", "1b", "2b", "3b", "4b", "5b"],
                "rows": [
                    ["Chlapci", "6,0m a menej", "6,01–7,5m", "7,51–9,0m", "9,01–10,5m", "10,51m a viac"],
                    ["Dievčatá", "4,5m a menej", "4,51–5,8m", "5,81–7,0m", "7,01–8,3m", "8,31m a viac"],
                ],
            },
        },
        {
            "title": "Ľah-sed",
            "members": 2,
            "max_per_member": 5,
            "columns": ["Počet (za 1 min)", "Body"],
            "legend": ["Čím viac opakovaní za 1 minútu, tým viac bodov. Max. 5 bodov na člena, max. 10 za tím."],
            "scoring_table": {
                "headers": ["Body", "1b", "2b", "3b", "4b", "5b"],
                "rows": [
                    ["Chlapci", "25 a menej", "26–33", "34–41", "42–49", "50 a viac"],
                    ["Dievčatá", "20 a menej", "21–27", "28–35", "36–43", "44 a viac"],
                ],
            },
        },
    ]

    FILL_SCHOOL = PatternFill("solid", fgColor="D6E4F0")
    FONT_SMALL = Font(name="Calibri", size=9)
    FONT_SMALL_BOLD = Font(name="Calibri", size=9, bold=True)

    for sport in SPORT_SHEETS:
        ws_s = wb.create_sheet(title=sport["title"])

        # Title
        num_cols = len(["Škola", "Člen"] + sport["columns"])
        last_col = get_column_letter(num_cols)
        ws_s.merge_cells(f"A1:{last_col}1")
        c = ws_s.cell(row=1, column=1)
        c.value = f"MASÉR V AKCII – {sport['title']}"
        c.font = Font(name="Calibri", size=14, bold=True, color="1F4E79")
        c.alignment = ALIGN_CENTER
        ws_s.row_dimensions[1].height = 22

        # Scoring table first (at top, so it's visible when filling in)
        row = 3
        if sport.get("scoring_table"):
            st = sport["scoring_table"]
            for ci, hdr in enumerate(st["headers"], 1):
                cell = ws_s.cell(row=row, column=ci, value=hdr)
                cell.font = FONT_SMALL_BOLD
                cell.fill = FILL_HEADER
                cell.font = Font(name="Calibri", size=9, bold=True, color="FFFFFF")
                cell.alignment = ALIGN_CENTER
                cell.border = THIN_BORDER
            row += 1
            for data_row in st["rows"]:
                for ci, val in enumerate(data_row, 1):
                    cell = ws_s.cell(row=row, column=ci, value=val)
                    cell.font = FONT_SMALL
                    cell.alignment = ALIGN_CENTER
                    cell.border = THIN_BORDER
                    if ci == 1:
                        cell.font = FONT_SMALL_BOLD
                        cell.alignment = ALIGN_LEFT
                ws_s.row_dimensions[row].height = 15
                row += 1
            row += 1
        else:
            for line in sport["legend"]:
                ws_s.merge_cells(start_row=row, start_column=1, end_row=row, end_column=num_cols)
                ws_s.cell(row=row, column=1, value=line).font = FONT_SMALL
                ws_s.cell(row=row, column=1).alignment = ALIGN_LEFT
                ws_s.row_dimensions[row].height = 14
                row += 1
            row += 1

        # Table header
        headers = ["Škola", "Člen"] + sport["columns"]
        for ci, hdr in enumerate(headers, 1):
            cell = ws_s.cell(row=row, column=ci, value=hdr)
            cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
            cell.fill = FILL_HEADER
            cell.alignment = ALIGN_CENTER
            cell.border = THIN_BORDER
        ws_s.row_dimensions[row].height = 18
        row += 1

        # Data rows
        for school in SCHOOLS:
            ws_s.merge_cells(start_row=row, start_column=1, end_row=row + 1, end_column=1)
            cell = ws_s.cell(row=row, column=1, value=school)
            cell.font = FONT_SMALL_BOLD
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            cell.border = THIN_BORDER
            cell.fill = FILL_SCHOOL
            ws_s.cell(row=row + 1, column=1).border = THIN_BORDER
            ws_s.cell(row=row + 1, column=1).fill = FILL_SCHOOL

            for member in range(sport["members"]):
                r = row + member
                ws_s.cell(row=r, column=2, value=f"Člen {member + 1}").font = FONT_SMALL
                ws_s.cell(row=r, column=2).alignment = ALIGN_CENTER
                ws_s.cell(row=r, column=2).border = THIN_BORDER
                for ci in range(3, len(headers) + 1):
                    cell = ws_s.cell(row=r, column=ci)
                    cell.border = THIN_BORDER
                    cell.alignment = ALIGN_CENTER
                ws_s.row_dimensions[r].height = 20

            # Team total row
            r = row + sport["members"]
            ws_s.merge_cells(start_row=r, start_column=1, end_row=r, end_column=len(headers) - 1)
            cell = ws_s.cell(row=r, column=1, value="Spolu za tím")
            cell.font = FONT_SMALL_BOLD
            cell.alignment = Alignment(horizontal="right", vertical="center")
            cell.border = THIN_BORDER
            for ci in range(2, len(headers) + 1):
                ws_s.cell(row=r, column=ci).border = THIN_BORDER
            total_cell = ws_s.cell(row=r, column=len(headers))
            total_cell.border = THIN_BORDER
            total_cell.alignment = ALIGN_CENTER
            total_cell.font = FONT_SMALL_BOLD
            ws_s.row_dimensions[r].height = 16

            row += sport["members"] + 1

        # Column widths
        ws_s.column_dimensions["A"].width = 22
        ws_s.column_dimensions["B"].width = 8
        for ci in range(3, len(headers) + 1):
            ws_s.column_dimensions[get_column_letter(ci)].width = 12

        # Print setup — landscape A4, fit to 1 page
        ws_s.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
            fitToPage=True)
        ws_s.page_setup.fitToWidth = 1
        ws_s.page_setup.fitToHeight = 1
        ws_s.page_setup.orientation = "landscape"
        ws_s.page_setup.paperSize = ws_s.PAPERSIZE_A4
        ws_s.page_margins.left = 0.4
        ws_s.page_margins.right = 0.4
        ws_s.page_margins.top = 0.3
        ws_s.page_margins.bottom = 0.3

    # ═══ TEST SCORING SHEET ═══
    ws_test = wb.create_sheet(title="Test")

    ws_test.merge_cells("A1:C1")
    c = ws_test.cell(row=1, column=1)
    c.value = "MASÉR V AKCII – Test"
    c.font = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    c.alignment = ALIGN_CENTER
    ws_test.row_dimensions[1].height = 30

    row = 3
    for ci, hdr in enumerate(["Škola", "Body"], 1):
        cell = ws_test.cell(row=row, column=ci, value=hdr)
        cell.font = FONT_HEADER
        cell.fill = FILL_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws_test.row_dimensions[row].height = ROW_HEIGHT_HEADER
    row += 1

    for school in SCHOOLS:
        cell = ws_test.cell(row=row, column=1, value=school)
        cell.font = FONT_NORMAL
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        cell.fill = FILL_SCHOOL
        ws_test.cell(row=row, column=2).border = THIN_BORDER
        ws_test.cell(row=row, column=2).alignment = ALIGN_CENTER
        ws_test.row_dimensions[row].height = 28
        row += 1

    ws_test.column_dimensions["A"].width = 28
    ws_test.column_dimensions["B"].width = 14

    ws_test.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(
        fitToPage=True)
    ws_test.page_setup.fitToWidth = 1
    ws_test.page_setup.fitToHeight = 1
    ws_test.page_setup.orientation = "portrait"
    ws_test.page_setup.paperSize = ws_test.PAPERSIZE_A4
    ws_test.page_margins.left = 0.5
    ws_test.page_margins.right = 0.5
    ws_test.page_margins.top = 0.4
    ws_test.page_margins.bottom = 0.4

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
